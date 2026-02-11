from __future__ import annotations

import csv
from io import BytesIO

from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.utils import timezone
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.http import require_GET, require_POST

from .models import HistoriqueRetentionSetting, HistoriqueUser
from .services import purge_old_history


def _session_role(request) -> str:
    user = request.session.get("user") or {}
    if isinstance(user, dict):
        return str(user.get("role") or "")
    return ""


def _is_admin(request) -> bool:
    return _session_role(request) == "Administrateur"


@require_GET
def historique_fragment(request):
    if not _is_admin(request):
        return HttpResponse("Accès refusé.", status=403)

    retention_years = HistoriqueRetentionSetting.get_value()
    current_year = timezone.localdate().year
    years_kept = ", ".join(str(y) for y in range(current_year - retention_years + 1, current_year + 1))

    history_rows = (
        HistoriqueUser.objects.select_related("user")
        .all()
        .order_by("-date_action", "-heure_action")[:300]
    )

    return render(
        request,
        "app_historique/fragment_historique.html",
        {
            "history_rows": history_rows,
            "retention_years": retention_years,
            "current_year": current_year,
            "years_kept": years_kept,
        },
    )


@csrf_protect
@require_POST
def historique_settings(request):
    if not _is_admin(request):
        return JsonResponse({"ok": False, "message": "Accès refusé."}, status=403)

    if "btn_save_retention" in request.POST:
        try:
            value = int(request.POST.get("retention_years") or "2")
        except ValueError:
            value = 2
        value = max(1, min(10, value))

        setting, _ = HistoriqueRetentionSetting.objects.get_or_create(id=1)
        setting.retention_years = value
        setting.save()

        deleted = purge_old_history()
        return JsonResponse({"ok": True, "deleted": deleted, "message": f"Rétention enregistrée ({value} an(s))."})

    if "btn_purge_history" in request.POST:
        deleted = purge_old_history()
        return JsonResponse({"ok": True, "deleted": deleted, "message": "Purge effectuée."})

    return JsonResponse({"ok": False, "message": "Action inconnue."}, status=400)


def _rows_for_export():
    return (
        HistoriqueUser.objects.select_related("user")
        .all()
        .order_by("-date_action", "-heure_action")
    )


def _export_filename(ext: str) -> str:
    ts = timezone.localtime().strftime("%Y%m%d_%H%M%S")
    return f"historique_utilisateur_{ts}.{ext}"


@require_GET
def export_historique_csv(request):
    if not _is_admin(request):
        return HttpResponse("Accès refusé.", status=403)

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{_export_filename("csv")}"'

    writer = csv.writer(response)
    writer.writerow(["DATE", "HEURE", "UTILISATEUR", "ACTION", "APP", "MODELE", "OBJET", "URL", "METHODE", "IP"])

    for h in _rows_for_export():
        writer.writerow([
            h.date_action,
            h.heure_action,
            getattr(h.user, "email", "") if h.user else "",
            h.action,
            h.app_label,
            h.model_name,
            h.object_repr,
            h.url,
            h.method,
            h.ip,
        ])
    return response


@require_GET
def export_historique_excel(request):
    if not _is_admin(request):
        return HttpResponse("Accès refusé.", status=403)

    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError:
        return HttpResponse(
            "Export Excel indisponible: installez la dépendance: pip install openpyxl",
            status=501,
            content_type="text/plain; charset=utf-8",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Historique"

    headers = ["DATE", "HEURE", "UTILISATEUR", "ACTION", "APP", "MODELE", "OBJET", "URL", "METHODE", "IP"]
    ws.append(headers)

    for h in _rows_for_export():
        ws.append([
            str(h.date_action),
            str(h.heure_action),
            getattr(h.user, "email", "") if h.user else "",
            h.action,
            h.app_label,
            h.model_name,
            h.object_repr,
            h.url,
            h.method,
            h.ip,
        ])

    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(60, len(header) + 10))

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    response = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{_export_filename("xlsx")}"'
    return response


@require_GET
def export_historique_pdf(request):
    if not _is_admin(request):
        return HttpResponse("Accès refusé.", status=403)

    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
    except ModuleNotFoundError:
        return HttpResponse(
            "Export PDF indisponible: installez la dépendance: pip install reportlab",
            status=501,
            content_type="text/plain; charset=utf-8",
        )

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=landscape(A4))
    width, height = landscape(A4)

    c.setFont("Helvetica-Bold", 14)
    c.drawString(15 * mm, height - 15 * mm, "Historique Utilisateur")
    c.setFont("Helvetica", 9)
    c.drawString(15 * mm, height - 21 * mm, f"Généré le {timezone.localtime().strftime('%d/%m/%Y %H:%M:%S')}")

    y = height - 30 * mm
    line_h = 6.5 * mm

    headers = ["DATE", "HEURE", "UTILISATEUR", "ACTION", "APP", "MODELE", "OBJET", "URL"]
    cols = [22*mm, 18*mm, 55*mm, 20*mm, 25*mm, 30*mm, 65*mm, width - (15*mm + 22*mm+18*mm+55*mm+20*mm+25*mm+30*mm+65*mm) - 15*mm]
    x0 = 15 * mm

    def draw_row(values, y_pos, bold=False):
        c.setFont("Helvetica-Bold" if bold else "Helvetica", 8.5)
        x = x0
        for val, w in zip(values, cols):
            txt = "" if val is None else str(val)
            if len(txt) > 120:
                txt = txt[:117] + "..."
            c.drawString(x, y_pos, txt)
            x += w

    draw_row(headers, y, bold=True)
    y -= line_h
    c.line(x0, y + 2*mm, width - 15*mm, y + 2*mm)

    for h in _rows_for_export():
        if y < 15 * mm:
            c.showPage()
            y = height - 15 * mm
        draw_row([
            h.date_action,
            h.heure_action,
            getattr(h.user, "email", "") if h.user else "",
            h.action,
            h.app_label,
            h.model_name,
            h.object_repr,
            h.url,
        ], y, bold=False)
        y -= line_h

    c.save()
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{_export_filename("pdf")}"'
    return response
